import json
import os
import openpyxl

class FileManager:
    """Manages file operations for data storage."""
    @staticmethod
    def ensure_directory_exists(directory_path):
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)

    staticmethod
    def save_to_json(data, filename, directory_path):
        FileManager.ensure_directory_exists(directory_path)  # Make sure the directory exists
        full_path = os.path.join(directory_path, filename)
        with open(full_path, 'w') as f:
            json.dump(data, f, indent=4)

    @staticmethod
    def save_to_csv(dataframe, filename, directory_path):
        FileManager.ensure_directory_exists(directory_path)
        full_path = os.path.join(directory_path, filename)
        dataframe.to_csv(full_path, index=False)

    @staticmethod
    def save_stat_to_workbook(dataframe, filename, directory_path):
        """ saves the stats dataframe format to the file destination"""
        # convert from dataframe to excel, using the stat xlsm template
        # Load the target format workbook - ensure template workbook is in templates folder
        # dont load work book!!! Instead we use a workbook as template and create a copy in storage and load that
        # Load the target format workbook
        workbook = openpyxl.load_workbook("Psyc_MonthlyStats_Aug 2023_JJ.xlsm")
        sheet = workbook.active

        # Find the row with the label "KEY IN YOUR INPUT FROM HERE ↓" - oad the Excel workbook with the given filename and then select the active (or first) sheet within that workbook
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1, values_only=True), start=1):
            if row[0] == "KEY IN YOUR INPUT FROM HERE ↓":
                start_row = row_num + 1
                break

        # Insert the transformed data into the target format starting from the identified row
        for index, row in dataframe.iterrows():
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=start_row + index, column=col_num, value=value)
        
        FileManager.ensure_directory_exists(directory_path)
        full_path = os.path.join(directory_path, filename)
        dataframe.to_excel()